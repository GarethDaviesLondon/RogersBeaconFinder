#Analysis for Xodus Salesforce.com Opportunity History Export
#
#   (c) Klynetic Innovation, November 2021
#   Author: Gareth Davies
#

#   
##################################################################################

import openpyxl
import tkinter as tk
from datetime import datetime, timedelta
from tkinter import filedialog
from openpyxl.worksheet import worksheet

def wo (op,st):
        try: 
            op.write (st)
        except:
            op.write ("ERROR")
def nl ():
      wo(op,"\n")

#This is required for the dialog boxes
root = tk.Tk

sfdata = filedialog.askopenfilename()
wb_obj= openpyxl.load_workbook(sfdata)
op =  open("beacons.h","w")

wo(op,"struct Beacon {")
wo(op,"   const char* Name;")
wo(op,"	double Mhz;")
wo(op,"	const char* Locator;")

wo(op,"};")
nl()
wo(op,"struct BandBeaconList {")
wo(op,"    const char* Name;")
wo(op,"    Beacon[] BeaconList;")
wo(op,"};")
nl()

for ws in wb_obj.sheetnames:

    ###### INGEST THE DATA
    sheet = wb_obj[ws]
    col_names=[]
    data = {}
    for column in sheet.iter_cols(1,sheet.max_column):
        col_names.append(column[0].value)
        data[column[0].value] = []
    numberofrows=0

    for i, row in enumerate(sheet.iter_rows()) :
        if i != 0:
            colindex =0
            for col in col_names:
                data[col].append(row[colindex].value)
                colindex+=1
        numberofrows=i

    ######## END OF INGESTING DATA

    ######## Set output file name

    rowcount=0

    wo(op,"Beacon B")
    wo(op,ws)
    wo(op,"[]=")
    nl()
    wo(op,"{")
    while (rowcount<numberofrows):
        #initiate each row for analysis
       
       wo(op,"{\"")
       wo(op, data['CALL'][rowcount])
       wo(op,"\",")
       freq=str(data["QRG"][rowcount])
       wo(op,freq)
        #op.write (data["QRG"][rowcount].toString())
       wo(op,",\"")
       wo(op,data["QTH"][rowcount])
       wo(op,"\"}")
       
       rowcount=rowcount+1
       if (rowcount<numberofrows) :     wo (op, ",")
       nl()
    wo(op,"};")
    nl()

wo(op,"BandBeaconList Bandbeacon[]={")
for ws in wb_obj.sheetnames:
    wo(op,"B")
    wo(op,ws)
    wo(op,",")
wo(op,"};")

op.close()

"""
    op.write ("Band =")
    op.write(ws)
    op.write ("{")
    while (rowcount<numberofrows):
        #initiate each row for analysis
        wo(op,"x")
        op.write ("{\"")
        op.write( data['CALL'][rowcount])
        op.write("\",\"")
        freq=str(data["QRG"][rowcount])
        op.write (freq)
        #op.write (data["QRG"][rowcount].toString())
        op.write ("\",\"")
        op.write (data["QTH"][rowcount])
        op.write("\"}")

        rowcount=rowcount+1
        if (rowcount<numberofrows) :     op.write( ",")#
        op.write ("};")
    op.close()
"""

